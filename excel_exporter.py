"""
Excel 导出模块 - 生成明细表 + 汇总表
"""

import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


def export_to_excel(expenses: list[dict], output_path: str = None) -> str:
    """
    导出 Excel，包含两个 Sheet：
    Sheet1: 明细表（日期/报销人/品类/金额/备注/共计）
    Sheet2: 汇总表（按人员统计 + 按品类统计 + 占比）
    """
    if output_path is None:
        if getattr(sys, 'frozen', False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(base, f"报销统计_{timestamp}.xlsx")

    wb = Workbook()

    # ========== Sheet 1: 明细表 ==========
    ws1 = wb.active
    ws1.title = "报销明细"

    _write_detail_sheet(ws1, expenses)

    # ========== Sheet 2: 汇总表 ==========
    ws2 = wb.create_sheet("汇总统计")
    _write_summary_sheet(ws2, expenses)

    wb.save(output_path)
    return output_path


def _create_header_style():
    """表头样式"""
    return {
        "font": Font(name="微软雅黑", size=11, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    }


def _create_cell_style():
    """普通单元格样式"""
    return {
        "font": Font(name="微软雅黑", size=10),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    }


def _apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def _write_detail_sheet(ws, expenses):
    """写入明细表"""
    headers = ["序号", "日期", "报销人", "品类", "金额（元）", "备注"]
    header_style = _create_header_style()
    cell_style = _create_cell_style()

    # 表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_style(cell, header_style)

    # 数据行
    total_amount = 0
    for i, exp in enumerate(expenses, 1):
        row_data = [
            i,
            exp.get("date", ""),
            exp.get("person", ""),
            exp.get("category", ""),
            exp.get("amount", 0),
            exp.get("remark", "")
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            _apply_style(cell, cell_style)
            if col == 5:
                cell.number_format = '#,##0.00'
        total_amount += exp.get("amount", 0)

    # 合计行
    total_row = len(expenses) + 2
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    total_label = ws.cell(row=total_row, column=1, value="合计")
    total_label.font = Font(name="微软雅黑", size=11, bold=True)
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_label.border = cell_style["border"]
    # 给合并区域的每个单元格加边框
    for c in range(2, 5):
        ws.cell(row=total_row, column=c).border = cell_style["border"]

    total_cell = ws.cell(row=total_row, column=5, value=round(total_amount, 2))
    total_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FF0000")
    total_cell.alignment = cell_style["alignment"]
    total_cell.border = cell_style["border"]
    total_cell.number_format = '#,##0.00'

    ws.cell(row=total_row, column=6).border = cell_style["border"]

    # 列宽
    col_widths = [6, 14, 12, 14, 16, 25]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_summary_sheet(ws, expenses):
    """写入汇总表"""
    header_style = _create_header_style()
    cell_style = _create_cell_style()

    total_amount = sum(e.get("amount", 0) for e in expenses)

    # ---- 按人员统计 ----
    person_header_style = _create_header_style()
    person_header_style["fill"] = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="按人员统计")
    title_cell.font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    p_headers = ["报销人", "笔数", "合计金额（元）", "占比"]
    for col, h in enumerate(p_headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        _apply_style(cell, person_header_style)

    person_stats = {}
    for e in expenses:
        p = e.get("person", "未知")
        if p not in person_stats:
            person_stats[p] = {"count": 0, "total": 0}
        person_stats[p]["count"] += 1
        person_stats[p]["total"] += e.get("amount", 0)

    person_sorted = sorted(person_stats.items(), key=lambda x: x[1]["total"], reverse=True)

    for i, (person, stat) in enumerate(person_sorted, 3):
        pct = (stat["total"] / total_amount * 100) if total_amount > 0 else 0
        row_data = [person, stat["count"], round(stat["total"], 2), f"{pct:.1f}%"]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            _apply_style(cell, cell_style)
            if col == 3:
                cell.number_format = '#,##0.00'

    # ---- 按品类统计 ----
    cat_start_row = len(person_sorted) + 4
    cat_header_style = _create_header_style()
    cat_header_style["fill"] = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    ws.merge_cells(start_row=cat_start_row - 1, start_column=1, end_row=cat_start_row - 1, end_column=4)
    cat_title = ws.cell(row=cat_start_row - 1, column=1, value="按品类统计")
    cat_title.font = Font(name="微软雅黑", size=13, bold=True, color="375623")
    cat_title.alignment = Alignment(horizontal="center", vertical="center")

    c_headers = ["品类", "笔数", "合计金额（元）", "占比"]
    for col, h in enumerate(c_headers, 1):
        cell = ws.cell(row=cat_start_row, column=col, value=h)
        _apply_style(cell, cat_header_style)

    cat_stats = {}
    for e in expenses:
        c = e.get("category", "其他")
        if c not in cat_stats:
            cat_stats[c] = {"count": 0, "total": 0}
        cat_stats[c]["count"] += 1
        cat_stats[c]["total"] += e.get("amount", 0)

    cat_sorted = sorted(cat_stats.items(), key=lambda x: x[1]["total"], reverse=True)

    for i, (cat, stat) in enumerate(cat_sorted, cat_start_row + 1):
        pct = (stat["total"] / total_amount * 100) if total_amount > 0 else 0
        row_data = [cat, stat["count"], round(stat["total"], 2), f"{pct:.1f}%"]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            _apply_style(cell, cell_style)
            if col == 3:
                cell.number_format = '#,##0.00'

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
